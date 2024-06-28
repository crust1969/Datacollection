import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import os
import tempfile

# Funktion zum Schreiben von Daten in die Excel-Datei
def write_to_excel(file_path, data):
    # Laden Sie die bestehende Excel-Datei
    book = load_workbook(file_path)
    writer = pd.ExcelWriter(file_path, engine='openpyxl')
    writer.book = book

    # Daten in ein DataFrame konvertieren
    df = pd.DataFrame(data, index=[0])

    # Daten in das Blatt schreiben
    df.to_excel(writer, sheet_name='Tabelle1', index=False, header=False, startrow=writer.sheets['Tabelle1'].max_row)

    # Excel-Datei speichern
    writer.save()
    writer.close()  # Close the writer after saving

# Hauptfunktion der Streamlit-App
def main():
    st.title("Daten in Excel-Tabelle schreiben")

    # Dateiupload für die Excel-Datei
    uploaded_file = st.file_uploader("Wählen Sie eine Excel-Datei aus", type=["xlsx"])

    if uploaded_file:
        # Save the uploaded file to a temporary location
        temp_file = tempfile.NamedTemporaryFile(delete=False)
        temp_file.write(uploaded_file.read())
        file_path = temp_file.name

        # Anzeigen der hochgeladenen Datei
        st.write("Hochgeladene Datei:", uploaded_file.name)

        # Dateneingabe für die Excel-Tabelle
        with st.form("data_form"):
            col1, col2 = st.columns(2)
            with col1:
                name = st.text_input("Name")
                age = st.number_input("Alter", min_value=0, max_value=120, step=1)
            with col2:
                gender = st.selectbox("Geschlecht", ["Männlich", "Weiblich", "Andere"])
                country = st.text_input("Land")

            submitted = st.form_submit_button("Speichern")

        if submitted:
            # Daten als Dictionary speichern
            data = {
                "Name": name,
                "Alter": age,
                "Geschlecht": gender,
                "Land": country
            }

            # Schreiben der Daten in die Excel-Datei (Übergabe des Dateipfads)
            write_to_excel(file_path, data)
            st.success("Daten wurden erfolgreich gespeichert!")

if __name__ == "__main__":
    main()
